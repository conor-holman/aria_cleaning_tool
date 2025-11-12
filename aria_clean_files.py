import pandas as pd
import numpy as np
import os
from tkinter import Tk, filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from typing import Tuple, List, Optional
import os
import unicodedata  # <-- Add this import at the top of your file



# Define essential columns that cannot be empty
ESSENTIAL_COLUMNS = [
    'PERSON_ID', 'FIRST_NAME', 'SURNAME', 'EMAIL', 'COMPANY',
    'DUTY_ENTITLEMENT', 'FIP_ENTITLEMENT', 'PALS_ENTITLEMENT',
    'BLUE_VOUCHERS', 'RESULT'
]

# Define columns and their desired string formatting
STRING_FORMATTING_RULES = {
    'FIRST_NAME': str.capitalize,
    'SURNAME': str.capitalize,
    'COMPANY': str.capitalize,
    'MANAGER_NAME': str.capitalize,
    'EMAIL': str.lower,
    'MANAGER_EMAIL': str.lower,
    'RESULT': str.upper
}

# Define valid values for specific columns
VALID_RESULT_VALUES = ['INS', 'DEL', 'UPD']
VALID_ENTITLEMENT_VALUES = ['Y', 'N']

def select_file() -> Optional[str]:
    """Opens a dialog for the user to select a CSV file."""
    # The Tk root is now managed by main()
    filepath = filedialog.askopenfilename(
        title="Select a CSV file",
        filetypes=[("CSV files", "*.csv")]
    )
    return filepath if filepath else None

def load_data(filepath: str) -> Optional[pd.DataFrame]:
    """
    Loads data from a CSV file, trying UTF-8 and then ISO-8851-1 encoding.
    Also trims whitespace from all string cells.
    """
    try:
        df = pd.read_csv(filepath, encoding='utf-8')
    except UnicodeDecodeError:
        print("UTF-8 decoding failed, trying ISO-8859-1...")
        df = pd.read_csv(filepath, encoding='iso-8859-1')
    except FileNotFoundError:
        print(f"Error: File not found at {filepath}")
        return None

    # Trim whitespace from all string columns
    df = df.apply(lambda x: x.str.strip() if x.dtype == "object" else x)
    return df

def convert_accented_text(text: str) -> str:
    """
    Converts accented characters to their ASCII equivalents using built-ins.
    e.g., 'René' -> 'Rene'
    """

    if not isinstance(text, str):
        return text
    
    # 1. Normalize using NFKD to separate base characters from diacritics
    normalized = unicodedata.normalize('NFKD', text)
    # 2. Encode to ASCII, 'ignore' drops the non-ASCII diacritics
    ascii_bytes = normalized.encode('ascii', 'ignore')
    # 3. Decode back to a standard string
    return ascii_bytes.decode('utf-8')

def clean_and_validate(df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Cleans, validates, and formats the DataFrame.
    Returns a tuple of (cleaned_df, dropped_df).
    """
    df_copy = df.copy()
    df_copy['error_reason'] = ''

    # --- NEW: Parse dates for ALL rows first to allow validation ---
    df_copy['JOINING_DATE'] = pd.to_datetime(df_copy['JOINING_DATE'], errors='coerce', dayfirst=True)
    df_copy['LEAVING_DATE'] = pd.to_datetime(df_copy['LEAVING_DATE'], errors='coerce', dayfirst=True)

    # 1. Validate 'RESULT' column
    invalid_result_mask = ~df_copy['RESULT'].isin(VALID_RESULT_VALUES)
    df_copy.loc[invalid_result_mask, 'error_reason'] += 'Invalid RESULT value; (must be INS for Insert, DEL for Deletion, UPD for Updating)'

    # 2. Validate essential columns for missing values
    for col in ESSENTIAL_COLUMNS:
        if col in df_copy.columns:
            missing_mask = df_copy[col].isna()
            df_copy.loc[missing_mask, 'error_reason'] += f'Missing {col}; '

    # 3. Validate and format entitlement columns
    for col in ['DUTY_ENTITLEMENT', 'PALS_ENTITLEMENT', 'FIP_ENTITLEMENT']:
        if col in df_copy.columns:
            # Only apply .str.upper() to non-NaN values
            df_copy[col] = df_copy[col].apply(lambda x: x.upper() if isinstance(x, str) else x)
            invalid_entitlement_mask = ~df_copy[col].isin(VALID_ENTITLEMENT_VALUES) & df_copy[col].notna()
            df_copy.loc[invalid_entitlement_mask, 'error_reason'] += f'Invalid {col} (must be Y/N); '


    # 4. Validate 'BLUE_VOUCHERS' is numeric
    # Convert 'N' or 'n' to '0' before numeric conversion
    df_copy['BLUE_VOUCHERS'] = df_copy['BLUE_VOUCHERS'].replace(['N', 'n'], '0') 
    
    numeric_vouchers = pd.to_numeric(df_copy['BLUE_VOUCHERS'], errors='coerce')
    invalid_voucher_mask = numeric_vouchers.isna() & df_copy['BLUE_VOUCHERS'].notna()
    df_copy.loc[invalid_voucher_mask, 'error_reason'] += 'BLUE_VOUCHERS not numeric; '
    df_copy['BLUE_VOUCHERS'] = numeric_vouchers

    # --- 5. NEW: Validate Dates ---
    upd_mask = df_copy['RESULT'] == 'UPD'
    invalid_dates_mask = upd_mask & df_copy['JOINING_DATE'].isna() & df_copy['LEAVING_DATE'].isna()
    df_copy.loc[invalid_dates_mask, 'error_reason'] += 'UPD row must have at least one valid date; '

    # --- Separate dropped rows from the main dataframe ---
    dropped_mask = df_copy['error_reason'] != ''
    dropped_df = df_copy[dropped_mask]
    cleaned_df = df_copy[~dropped_mask].drop(columns=['error_reason'])

    if cleaned_df.empty:
        # Return early if no rows are clean
        return cleaned_df, dropped_df

    # --- Apply Formatting to Cleaned Data ---
    for col, formatting_func in STRING_FORMATTING_RULES.items():
        if col in cleaned_df.columns:
            
            # --- NEW ROBUST FORMATTING ---
            # This lambda function checks if the cell is a string.
            # If it is: normalize it, then apply the formatting function.
            # If it's not (e.g., it's NaN, None, or a number), leave it alone.
            def format_and_normalize(cell_value):
                if isinstance(cell_value, str):
                    normalized_val = convert_accented_text(cell_value)
                    return formatting_func(normalized_val)
                return cell_value

            cleaned_df[col] = cleaned_df[col].apply(format_and_normalize)

    # --- Handle Dates ---
    today = pd.to_datetime('today').normalize()

    # Adjust dates based on 'RESULT'
    ins_mask = cleaned_df['RESULT'] == 'INS'
    del_mask = cleaned_df['RESULT'] == 'DEL'

    # If RESULT is INS, adjust JOINING_DATE
    cleaned_df.loc[ins_mask, 'JOINING_DATE'] = cleaned_df.loc[ins_mask, 'JOINING_DATE'].fillna(today)
    cleaned_df.loc[ins_mask, 'LEAVING_DATE'] = pd.NaT

    # If RESULT is DEL, adjust LEAVING_DATE
    cleaned_df.loc[del_mask, 'LEAVING_DATE'] = cleaned_df.loc[del_mask, 'LEAVING_DATE'].fillna(today)
    cleaned_df.loc[del_mask, 'JOINING_DATE'] = pd.NaT
    
    # Convert both columns to string AT THE END
    cleaned_df['JOINING_DATE'] = cleaned_df['JOINING_DATE'].dt.strftime('%Y.%m.%d %H:%M:%S')
    cleaned_df['LEAVING_DATE'] = cleaned_df['LEAVING_DATE'].dt.strftime('%Y.%m.%d %H:%M:%S')

    # --- Final Touches ---
    datestamp = pd.Timestamp.now().strftime('%Y.%m.%d %H:%M:%S')
    cleaned_df['DATESTAMP'] = datestamp

    return cleaned_df, dropped_df

def save_results(cleaned_df: pd.DataFrame, dropped_df: pd.DataFrame, original_path: str) -> Tuple[str, Optional[str]]:
    """
    Saves the cleaned data and the highlighted dropped rows.
    Returns the paths to the saved files (cleaned_csv_path, dropped_excel_path).
    """
    base_name = os.path.splitext(os.path.basename(original_path))[0]
    output_dir = os.path.dirname(original_path)

    # --- Paths to be returned ---
    cleaned_csv_path = os.path.join(output_dir, f"{base_name}_cleaned.csv")
    dropped_excel_path = None # Default to None

    # Save cleaned data
    # Ensure date columns (now strings or NaT) are saved correctly
    cleaned_df_for_csv = cleaned_df.copy()
    cleaned_df_for_csv['JOINING_DATE'] = cleaned_df_for_csv['JOINING_DATE'].replace({pd.NaT: '', 'NaT': ''})
    cleaned_df_for_csv['LEAVING_DATE'] = cleaned_df_for_csv['LEAVING_DATE'].replace({pd.NaT: '', 'NaT': ''})
    
    cleaned_df_for_csv.to_csv(cleaned_csv_path, index=False)
    print(f"✅ Cleaned data saved to: {cleaned_csv_path}")

    # Save dropped rows if any exist
    if not dropped_df.empty:
        dropped_csv_path = os.path.join(output_dir, f"{base_name}_dropped.csv")
        dropped_excel_path = os.path.join(output_dir, f"{base_name}_dropped_highlighted.xlsx")
        
        dropped_df.to_csv(dropped_csv_path, index=False, date_format='%Y.%m.%d')
        print(f"❌ Dropped rows saved to: {dropped_csv_path}")

        # Create highlighted Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "Dropped Rows"
        
        # Write header
        headers = list(dropped_df.columns)
        ws.append(headers)
        
        # Define fill style
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

        # --- NEW HIGHLIGHTING LOGIC ---
        # Write data and highlight cells based on the 'error_reason' column
        for _, row in dropped_df.iterrows():
            # Convert row to list, handling NaT dates for Excel
            excel_row_data = []
            for item in row.tolist():
                if pd.isna(item): # Append None for NaT or NaN
                    excel_row_data.append(None) 
                else:
                    excel_row_data.append(item)
            ws.append(excel_row_data)
            excel_row = ws.max_row
            
            error_reason = row.get('error_reason', '')
            
            # Check for our new specific date error
            date_error = 'at least one valid date' in error_reason

            # Iterate through columns to decide on highlighting
            for col_idx, col_name in enumerate(headers, 1):
                cell_to_fill = ws.cell(row=excel_row, column=col_idx)
                
                # 1. Specific Date Error: Highlight both date columns
                if date_error and (col_name == 'JOINING_DATE' or col_name == 'LEAVING_DATE'):
                    cell_to_fill.fill = red_fill
                
                # 2. Missing Essential Columns: Check if 'Missing {col_name}' is in the error
                elif f'Missing {col_name}' in error_reason:
                    cell_to_fill.fill = red_fill
                
                # 3. Invalid Entitlements or RESULT
                elif f'Invalid {col_name}' in error_reason:
                    cell_to_fill.fill = red_fill
                
                # 4. Non-numeric Blue Vouchers
                elif col_name == 'BLUE_VOUCHERS' and 'BLUE_VOUCHERS not numeric' in error_reason:
                    cell_to_fill.fill = red_fill
        
        wb.save(dropped_excel_path)
        print(f"🎨 Highlighted dropped rows saved to: {dropped_excel_path}")
    
    return cleaned_csv_path, dropped_excel_path

def main():
    """Main function to run the data cleaning process."""
    
    # --- Create a single, hidden Tk root for all dialogs ---
    root = Tk()
    root.withdraw()

    filepath = select_file()
    if not filepath:
        print("No file selected. Exiting.")
        messagebox.showerror("Error", "No file selected. Exiting.")
        root.destroy()
        return

    print(f"Processing file: {filepath}")
    df = load_data(filepath)
    if df is None:
        messagebox.showerror("Error", f"Could not load file: {filepath}")
        root.destroy()
        return

    cleaned_df, dropped_df = clean_and_validate(df)

    # --- Summary for print and final pop-up ---
    total_rows = len(df)
    cleaned_rows = len(cleaned_df)
    dropped_rows = len(dropped_df)

    summary_text = (
        f"Total rows read: {total_rows}\n"
        f"Rows cleaned and saved: {cleaned_rows}\n"
        f"Rows dropped due to errors: {dropped_rows}\n"
    )
    
    print("\n--- Processing Summary ---")
    print(summary_text)
    print("------------------------\n")

    cleaned_path, dropped_path = save_results(cleaned_df, dropped_df, filepath)
    print("\nProcessing complete.")

    # --- Build and show the final summary pop-up ---
    
    # Create the detailed message for the pop-up
    popup_message = f"{summary_text}\n--- Files Saved ---\n\n"
    popup_message += f"Cleaned data:\n{cleaned_path}\n\n"
    
    if dropped_path:
        popup_message += f"Dropped rows (highlighted):\n{dropped_path}"
    else:
        popup_message += "No rows were dropped."

    messagebox.showinfo("Processing Complete", popup_message)
    
    # Clean up the Tk root
    root.destroy()

if __name__ == "__main__":
    main()